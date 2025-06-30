"""
Módulo para manipulação de arquivos Excel
=========================================

Funções para leitura de planilhas de colaboradores e feriados,
bem como exportação de escalas geradas.
"""

import pandas as pd
import streamlit as st
from datetime import datetime
import os
from io import BytesIO
from openpyxl.styles import Font, PatternFill

def ler_planilha_colaboradores(arquivo):
    """
    Lê a planilha de colaboradores e valida os dados.
    
    Args:
        arquivo: Arquivo uploader do Streamlit
        
    Returns:
        pandas.DataFrame: DataFrame com os dados dos colaboradores
        
    Raises:
        ValueError: Se a planilha não contém as colunas obrigatórias
    """
    try:
        df = pd.read_excel(arquivo)
        
        # Colunas obrigatórias
        colunas_obrigatorias = [
            'Nome', 'Cargo', 'Tipo_Escala', 'Turno'
        ]
        
        # Colunas opcionais
        colunas_opcionais = [
            'Atestados', 'Ferias', 'Escalas_Manuais', 'Ultimo_Plantao_Mes_Anterior', 'Ultimo_Domingo_Folga'
        ]
        
        # Verificar colunas obrigatórias
        colunas_faltantes = [col for col in colunas_obrigatorias if col not in df.columns]
        if colunas_faltantes:
            raise ValueError(f"Colunas obrigatórias não encontradas: {colunas_faltantes}")
        
        # Adicionar colunas opcionais se não existirem
        for coluna in colunas_opcionais:
            if coluna not in df.columns:
                df[coluna] = ''
        
        # Validar tipos de escala
        tipos_escala_validos = [
            # Escalas 44H
            'M44', 'T44', 'N44',
            # Escalas 40H
            'M40', 'T40', 'N40',
            # Escalas 6X1
            'M6X1', 'T6X1', 'N6X1',
            # Plantões
            'P_D', 'P_N', 'I_D', 'I_N'
        ]
        escalas_invalidas = df[~df['Tipo_Escala'].isin(tipos_escala_validos)]['Tipo_Escala'].unique()
        if len(escalas_invalidas) > 0:
            st.warning(f"Tipos de escala inválidos encontrados: {escalas_invalidas}")
        
        # Validar turnos
        turnos_validos = ['Manhã', 'Tarde', 'Noite', 'Dia']
        turnos_invalidos = df[~df['Turno'].isin(turnos_validos)]['Turno'].unique()
        if len(turnos_invalidos) > 0:
            st.warning(f"Turnos inválidos encontrados: {turnos_invalidos}")
        
        # Limpar dados - preservar formato original das datas
        df = df.fillna('')
        
        # Converter campos de data para string preservando formato original
        campos_data = ['Ultimo_Plantao_Mes_Anterior', 'Ultimo_Domingo_Folga']
        for campo in campos_data:
            if campo in df.columns:
                df[campo] = df[campo].astype(str).apply(lambda x: x.strip() if x != 'nan' else '')
        
        # Converter outros campos para string
        campos_texto = ['Nome', 'Cargo', 'Tipo_Escala', 'Turno', 'Atestados', 'Ferias', 'Escalas_Manuais']
        for campo in campos_texto:
            if campo in df.columns:
                df[campo] = df[campo].astype(str).apply(lambda x: x.strip() if x != 'nan' else '')
        
        return df
        
    except Exception as e:
        raise ValueError(f"Erro ao ler planilha de colaboradores: {str(e)}")

def ler_planilha_feriados(arquivo):
    """
    Lê a planilha de feriados e valida os dados.
    
    Args:
        arquivo: Arquivo uploader do Streamlit
        
    Returns:
        pandas.DataFrame: DataFrame com os dados dos feriados
        
    Raises:
        ValueError: Se a planilha não contém as colunas obrigatórias
    """
    try:
        df = pd.read_excel(arquivo)
        
        # Verificar colunas obrigatórias
        if 'Data' not in df.columns:
            raise ValueError("Coluna 'Data' não encontrada na planilha de feriados")
        
        # Adicionar coluna de descrição se não existir
        if 'Descricao' not in df.columns:
            df['Descricao'] = 'Feriado'
        
        # Converter coluna de data
        try:
            df['Data'] = pd.to_datetime(df['Data']).dt.date
        except:
            raise ValueError("Erro ao converter coluna 'Data' para formato de data")
        
        # Limpar dados
        df = df.fillna('')
        
        return df
        
    except Exception as e:
        raise ValueError(f"Erro ao ler planilha de feriados: {str(e)}")

def exportar_escala_excel(escala_completa, data_inicio, data_fim, nome_arquivo=None):
    """
    Exporta a escala completa para arquivo Excel.
    
    Args:
        escala_completa (pandas.DataFrame): DataFrame com a escala completa
        data_inicio (date): Data de início do período
        data_fim (date): Data de fim do período
        nome_arquivo (str, optional): Nome do arquivo de saída
        
    Returns:
        bytes: Conteúdo do arquivo Excel em bytes
    """
    try:
        # Criar nome do arquivo se não fornecido
        if nome_arquivo is None:
            nome_arquivo = f"escala_{data_inicio.strftime('%Y%m')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Criar writer do Excel
        with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
            # Escrever escala completa
            escala_completa.to_excel(writer, sheet_name='Escala_Completa', index=False)
            
            # TODO: Adicionar outras abas se necessário
            # - Resumo por colaborador
            # - Estatísticas
            # - Configurações utilizadas
        
        # Ler arquivo gerado e retornar bytes
        with open(nome_arquivo, 'rb') as f:
            conteudo = f.read()
        
        # Remover arquivo temporário
        os.remove(nome_arquivo)
        
        return conteudo
        
    except Exception as e:
        raise ValueError(f"Erro ao exportar escala para Excel: {str(e)}")

def criar_template_colaboradores():
    """
    Cria um template Excel para upload de colaboradores.
    
    Returns:
        Bytes do arquivo Excel
    """
    # Dados de exemplo
    dados_exemplo = [
        {
            'Nome': 'João Silva',
            'Cargo': 'Analista',
            'Tipo_Escala': 'M44',
            'Turno': 'Manhã',
            'Atestados': '',
            'Ferias': '',
            'Escalas_Manuais': '',
            'Ultimo_Plantao_Mes_Anterior': '',
            'Ultimo_Domingo_Folga': ''
        },
        {
            'Nome': 'Maria Santos',
            'Cargo': 'Técnico',
            'Tipo_Escala': 'T44',
            'Turno': 'Tarde',
            'Atestados': '15/06/2025',
            'Ferias': '',
            'Escalas_Manuais': '',
            'Ultimo_Plantao_Mes_Anterior': '',
            'Ultimo_Domingo_Folga': ''
        },
        {
            'Nome': 'Pedro Oliveira',
            'Cargo': 'Auxiliar',
            'Tipo_Escala': 'N6X1',
            'Turno': 'Noite',
            'Atestados': '',
            'Ferias': '10/06/2025-20/06/2025',
            'Escalas_Manuais': '',
            'Ultimo_Plantao_Mes_Anterior': '',
            'Ultimo_Domingo_Folga': '25/05/2025'
        },
        {
            'Nome': 'Ana Costa',
            'Cargo': 'Supervisor',
            'Tipo_Escala': 'P_D',
            'Turno': 'Dia',
            'Atestados': '',
            'Ferias': '',
            'Escalas_Manuais': '',
            'Ultimo_Plantao_Mes_Anterior': '30/05/2025',
            'Ultimo_Domingo_Folga': ''
        },
        {
            'Nome': 'Carlos Lima',
            'Cargo': 'Operador',
            'Tipo_Escala': 'I_N',
            'Turno': 'Noite',
            'Atestados': '20/06/2025',
            'Ferias': '',
            'Escalas_Manuais': '',
            'Ultimo_Plantao_Mes_Anterior': '29/05/2025',
            'Ultimo_Domingo_Folga': ''
        }
    ]
    
    # Criar DataFrame
    df = pd.DataFrame(dados_exemplo)
    
    # Criar arquivo Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Colaboradores', index=False)
        
        # Obter a planilha para formatação
        worksheet = writer.sheets['Colaboradores']
        
        # Formatar cabeçalhos
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    output.seek(0)
    return output.getvalue()

def criar_template_feriados():
    """
    Cria um template da planilha de feriados.
    
    Returns:
        pandas.DataFrame: DataFrame template
    """
    template = pd.DataFrame({
        'Data': ['2025-06-12', '2025-06-19'],
        'Descricao': ['Corpus Christi', 'São João']
    })
    
    return template 